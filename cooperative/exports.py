"""
PDF / Excel export helpers.
Uses reportlab (PDF) and openpyxl (Excel) — both already in requirements.txt
but previously unused anywhere in the project.
"""
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import CoopSettings


def _pdf_response(filename):
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _excel_response(filename):
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def savings_statement_pdf(member):
    """A member's savings statement: running balance ledger + summary."""
    coop = CoopSettings.get_settings()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(coop.coop_name, styles["Title"]),
        Paragraph("Savings Statement", styles["Heading2"]),
        Spacer(1, 6),
        Paragraph(f"Member: {member.full_name} ({member.member_id})", styles["Normal"]),
        Paragraph(f"Generated: {timezone.now().strftime('%d %b %Y %H:%M')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    records = member.savings_records.order_by("date", "created_at")
    data = [["Date", "Transaction", "Type", "Amount (₦)", "Balance After (₦)"]]
    for r in records:
        data.append([
            r.date.strftime("%d %b %Y"), r.transaction_id, r.get_transaction_type_display(),
            f"{r.amount:,.2f}", f"{r.balance_after:,.2f}",
        ])
    if len(data) == 1:
        data.append(["—", "No savings records yet", "", "", ""])

    table = Table(data, repeatRows=1, colWidths=[65, 90, 90, 90, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#225c3a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))
    elements.append(Paragraph(f"Current balance: ₦{member.total_savings:,.2f}", styles["Heading3"]))

    doc.build(elements)
    buf.seek(0)
    resp = _pdf_response(f"savings-statement-{member.member_id}.pdf")
    resp.write(buf.read())
    return resp


def _single_sheet_excel(sheet_title, headers, rows, column_widths, filename):
    """Shared builder for the three single-table ledger exports below."""
    wb = Workbook()
    header_fill = PatternFill(start_color="225C3A", end_color="225C3A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    for col, width in zip("ABCDEFGHIJ", column_widths):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = _excel_response(filename)
    resp.write(buf.read())
    return resp


def member_register_excel(members):
    """Full member register: identity, contact, status, and savings balance."""
    rows = [
        [
            m.member_id, m.full_name, m.user.email, m.phone, m.get_status_display(),
            m.get_role_display(), float(m.total_savings), m.join_date.strftime("%Y-%m-%d"),
        ]
        for m in members
    ]
    return _single_sheet_excel(
        "Member Register",
        ["Member ID", "Full Name", "Email", "Phone", "Status", "Role", "Savings Balance (₦)", "Joined"],
        rows,
        [14, 24, 26, 16, 12, 12, 18, 14],
        f"member-register-{timezone.now().strftime('%Y%m%d')}.xlsx",
    )


def savings_ledger_excel(savings_records):
    """Full savings ledger: every deposit/withdrawal/dividend across all members."""
    rows = [
        [
            s.transaction_id, s.member.member_id, s.member.full_name,
            s.get_transaction_type_display(), float(s.amount), float(s.balance_after),
            s.date.strftime("%Y-%m-%d"), s.description,
        ]
        for s in savings_records
    ]
    return _single_sheet_excel(
        "Savings Ledger",
        ["Transaction ID", "Member ID", "Member Name", "Type", "Amount (₦)",
         "Balance After (₦)", "Date", "Description"],
        rows,
        [18, 12, 24, 14, 14, 16, 12, 30],
        f"savings-ledger-{timezone.now().strftime('%Y%m%d')}.xlsx",
    )


def loan_ledger_excel(loans):
    """Full loan ledger: every loan application regardless of status."""
    rows = [
        [
            l.loan_id, l.member.member_id, l.member.full_name, l.get_purpose_display(),
            float(l.amount_requested), float(l.amount_approved or 0),
            l.get_status_display(), float(l.outstanding_balance),
            l.risk_score or "—", l.application_date.strftime("%Y-%m-%d"),
        ]
        for l in loans
    ]
    return _single_sheet_excel(
        "Loan Ledger",
        ["Loan ID", "Member ID", "Member Name", "Purpose", "Requested (₦)", "Approved (₦)",
         "Status", "Outstanding (₦)", "ML Risk", "Applied"],
        rows,
        [14, 12, 24, 16, 16, 16, 12, 16, 10, 12],
        f"loan-ledger-{timezone.now().strftime('%Y%m%d')}.xlsx",
    )


def financial_report_excel(transactions, loans, savings_total, withdrawals_total, disbursed_total, repaid_total):
    """Cooperative-wide financial report as a multi-sheet Excel workbook."""
    wb = Workbook()
    header_fill = PatternFill(start_color="225C3A", end_color="225C3A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws.append(["CoopX Financial Report"])
    ws.append([f"Generated: {timezone.now().strftime('%d %b %Y %H:%M')}"])
    ws.append([])
    ws.append(["Metric", "Amount (₦)"])
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
    ws.append(["Total Savings Deposits", float(savings_total)])
    ws.append(["Total Withdrawals", float(withdrawals_total)])
    ws.append(["Net Savings Pool", float(savings_total) - float(withdrawals_total)])
    ws.append(["Total Loans Disbursed", float(disbursed_total)])
    ws.append(["Total Loans Repaid", float(repaid_total)])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # ── Transactions sheet ──
    ws2 = wb.create_sheet("Transactions")
    ws2.append(["Transaction ID", "Member", "Type", "Amount (₦)", "Status", "Date"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for t in transactions:
        ws2.append([
            t.transaction_id, t.member.full_name, t.get_transaction_type_display(),
            float(t.amount), t.get_status_display(), t.date.strftime("%Y-%m-%d %H:%M"),
        ])
    for col, width in zip("ABCDEF", [22, 24, 20, 14, 12, 18]):
        ws2.column_dimensions[col].width = width

    # ── Loans sheet ──
    ws3 = wb.create_sheet("Loans")
    ws3.append(["Loan ID", "Member", "Requested (₦)", "Approved (₦)", "Status", "Outstanding (₦)", "Applied"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for l in loans:
        ws3.append([
            l.loan_id, l.member.full_name, float(l.amount_requested),
            float(l.amount_approved or 0), l.get_status_display(),
            float(l.outstanding_balance), l.application_date.strftime("%Y-%m-%d"),
        ])
    for col, width in zip("ABCDEFG", [16, 24, 16, 16, 14, 18, 14]):
        ws3.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = _excel_response(f"financial-report-{timezone.now().strftime('%Y%m%d')}.xlsx")
    resp.write(buf.read())
    return resp
